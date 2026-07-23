import csv
import io
import json
import re
import uuid
from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import delete, select

from .config import get_settings
from .database import SessionLocal
from .embeddings import embed_texts
from .models import Document, DocumentChunk
from .storage import storage


def extract_text(data: bytes, filename: str, content_type: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf" or content_type == "application/pdf":
        return "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(data)).pages)
    if suffix == ".docx":
        document = DocxDocument(io.BytesIO(data))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# {sheet.title}")
            lines.extend(
                "\t".join("" if value is None else str(value) for value in row)
                for row in sheet.iter_rows(values_only=True)
            )
        return "\n".join(lines)
    if suffix == ".csv":
        decoded = data.decode("utf-8-sig", errors="replace")
        return "\n".join("\t".join(row) for row in csv.reader(io.StringIO(decoded)))
    if suffix == ".json":
        return json.dumps(json.loads(data), ensure_ascii=False, indent=2)
    return data.decode("utf-8", errors="replace")


def split_text(text: str, size: int, overlap: int) -> list[str]:
    normalized = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not normalized:
        return []
    chunks: list[str] = []
    start = 0
    while start < len(normalized):
        end = min(start + size, len(normalized))
        if end < len(normalized):
            boundary = max(
                normalized.rfind("\n", start + size // 2, end),
                normalized.rfind("。", start + size // 2, end),
            )
            if boundary > start:
                end = boundary + 1
        chunks.append(normalized[start:end].strip())
        if end >= len(normalized):
            break
        start = max(end - overlap, start + 1)
    return [chunk for chunk in chunks if chunk]


async def process_document(document_id: uuid.UUID | str) -> None:
    settings = get_settings()
    document_uuid = uuid.UUID(str(document_id))
    async with SessionLocal() as session:
        document = await session.scalar(select(Document).where(Document.id == document_uuid))
        if not document:
            return
        document.status = "processing"
        document.error_message = ""
        await session.commit()
        try:
            data = storage.read_bytes(document.object_key)
            text = extract_text(data, document.name, document.content_type)
            chunks = split_text(text, settings.chunk_size, settings.chunk_overlap)
            if not chunks:
                raise ValueError("文件中没有可索引的文本内容")
            embeddings = await embed_texts(chunks)
            await session.execute(delete(DocumentChunk).where(DocumentChunk.document_id == document.id))
            session.add_all(
                DocumentChunk(
                    knowledge_base_id=document.knowledge_base_id,
                    document_id=document.id,
                    ordinal=index,
                    content=chunk,
                    token_estimate=max(1, len(chunk) // 2),
                    embedding=embedding,
                    chunk_metadata={"filename": document.name},
                )
                for index, (chunk, embedding) in enumerate(zip(chunks, embeddings, strict=True))
            )
            document.chunk_count = len(chunks)
            document.status = "processed"
        except Exception as error:
            document.status = "error"
            document.error_message = str(error)[:1000]
        await session.commit()
